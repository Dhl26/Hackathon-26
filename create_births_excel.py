import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import numpy as np

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Registered Births"

# Function to predict future values using linear regression
def predict_value(data_points, years_ahead=1):
    """
    Predict future value using linear regression
    data_points: list of historical values (can contain 'N.A.' or None)
    years_ahead: number of years to predict ahead (1 for 2024, 2 for 2025)
    """
    # Filter out N.A. values and convert to numeric
    valid_data = []
    valid_indices = []
    for i, val in enumerate(data_points):
        if val != 'N.A.' and val is not None and val != '':
            try:
                valid_data.append(float(val))
                valid_indices.append(i)
            except (ValueError, TypeError):
                pass
    
    if len(valid_data) < 2:
        return None  # Not enough data for prediction
    
    # Perform linear regression
    x = np.array(valid_indices)
    y = np.array(valid_data)
    
    # Calculate slope and intercept
    coefficients = np.polyfit(x, y, 1)
    slope, intercept = coefficients
    
    # Predict for the next year(s)
    next_index = len(data_points) + years_ahead - 1
    prediction = slope * next_index + intercept
    
    # Round to nearest integer and ensure non-negative
    return max(0, int(round(prediction)))

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=10)
center_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 30
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
    ws.column_dimensions[col].width = 12

# Add headers
ws.merge_cells('A1:B1')
ws['A1'] = 'Sl. No.'
ws['A1'].fill = header_fill
ws['A1'].font = header_font
ws['A1'].alignment = center_alignment

ws.merge_cells('C1:N1')
ws['C1'] = 'Number of Registered Births'
ws['C1'].fill = header_fill
ws['C1'].font = header_font
ws['C1'].alignment = center_alignment

# Add year headers
years = ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024', '2025']
ws['A2'] = ''
ws['B2'] = 'Indian State/ Union Territory'
for idx, year in enumerate(years):
    ws.cell(row=2, column=idx+3).value = year

# Apply styling to row 2
for col in range(1, 15):
    cell = ws.cell(row=2, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment

# Data for India total
ws['A3'] = ''
ws['B3'] = 'India'
india_data = [23001523, 23136145, 22200991, 22104418, 23269383, 24820886, 24222444, 24201614, 25439164, 25207070]
# Predict 2024 using 2014-2023 data
india_2024 = predict_value(india_data, years_ahead=1)
india_data_with_2024 = india_data + [india_2024]
# Predict 2025 using 2014-2024 data
india_2025 = predict_value(india_data_with_2024, years_ahead=1)
india_data_full = india_data_with_2024 + [india_2025]

for idx, value in enumerate(india_data_full):
    ws.cell(row=3, column=idx+3).value = value
    ws.cell(row=3, column=idx+3).alignment = center_alignment

# States header
ws['A4'] = ''
ws['B4'] = 'States'
for col in range(1, 15):
    cell = ws.cell(row=4, column=col)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = center_alignment

# States data
states_data = [
    [1, 'Andhra Pradesh', 848883, 851499, 814754, 818165, 760716, 754839, 714017, 737189, 752403, 762093],
    [2, 'Arunachal Pradesh', 43246, 62058, 47963, 93539, 77249, 48837, 39493, 43644, 48711, 43397],
    [3, 'Assam', 723482, 1005881, 734454, 708744, 767856, 762318, 751694, 683173, 707036, 599739],
    [4, 'Bihar', 1694565, 1758356, 1697762, 2052273, 2241089, 2609167, 3044831, 3252977, 3071603, 2538619],
    [5, 'Chhattisgarh', 845337, 655847, 763878, 694030, 570056, 557723, 516918, 478572, 560641, 575467],
    [6, 'Goa', 21415, 21727, 21610, 20472, 20436, 19736, 18316, 15396, 17326, 17240],
    [7, 'Gujarat', 1205140, 1254838, 1257090, 1166848, 1168516, 1173759, 1103241, 1021362, 1126522, 1176320],
    [8, 'Haryana', 584739, 569340, 570747, 535198, 541405, 529952, 591914, 549829, 544109, 598417],
    [9, 'Himachal Pradesh', 106818, 115418, 101595, 101238, 97031, 94760, 90772, 92928, 90988, 94235],
    [10, 'Jharkhand', 642286, 691700, 698026, 699945, 669766, 719920, 646453, 554822, 702144, 974211],
    [11, 'Karnataka', 1087530, 1053248, 1107258, 1099019, 1028934, 1047877, 988143, 899065, 1037182, 1016541],
    [12, 'Kerala', 534458, 516013, 496292, 503507, 485174, 480113, 446891, 419787, 439742, 393231],
    [13, 'Madhya Pradesh', 1611269, 1519694, 1485779, 1486191, 1486794, 1602843, 1535628, 1577555, 1789365, 1988575],
    [14, 'Maharashtra', 1962601, 1941352, 1891451, 1809353, 1733501, 1747145, 1712003, 1718682, 1919023, 1805484],
    [15, 'Manipur', 58756, 59703, 'N.A.', 74179, 68733, 30120, 23895, 15809, 52833, 61203],
    [16, 'Meghalaya', 75223, 87242, 'N.A.', 75118, 86999, 130255, 112484, 102848, 127704, 127492],
    [17, 'Mizoram', 24516, 24786, 22497, 20964, 20984, 23808, 23320, 24314, 26049, 27180],
    [18, 'Nagaland', 58950, 79539, 72984, 69121, 47357, 77706, 46079, 84988, 80813, 85569],
    [19, 'Odisha', 800577, 779963, 772665, 692658, 676542, 675276, 671355, 646075, 676334, 653281],
    [20, 'Punjab', 463689, 436170, 463788, 417480, 403200, 391125, 361200, 368004, 400579, 440536],
    [21, 'Rajasthan', 1751191, 1769339, 1806205, 1689409, 1743150, 1795488, 1868383, 1900605, 1902140, 1932074],
    [22, 'Sikkim', 8057, 7878, 7132, 7124, 7035, 6639, 7189, 6212, 6177, 6506],
    [23, 'Tamil Nadu', 1206650, 1167506, 1062388, 948573, 915406, 941059, 934019, 913084, 936837, 902718],
    [24, 'Telangana', 569470, 612489, 624581, 617820, 652791, 841268, 669320, 611651, 703362, 652688],
    [25, 'Tripura', 45741, 51930, 43334, 79200, 73575, 72053, 66589, 67155, 70591, 88295],
    [26, 'Uttarakhand', 162869, 210445, 226175, 164167, 203822, 238061, 231689, 246220, 271320, 226085],
    [27, 'Uttar Pradesh', 3916825, 3881295, 3489890, 3545785, 4534084, 5131999, 4854098, 5075468, 5445903, 6070573],
    [28, 'West Bengal', 1330000, 1337146, 1415228, 1319441, 1473914, 1507813, 1475410, 1597267, 1358851, 1427576],
]

row_num = 5
for state in states_data:
    # Get historical data (columns 2 onwards are the year data)
    historical_data = state[2:]
    
    # Predict 2024 using 2014-2023 data
    pred_2024 = predict_value(historical_data, years_ahead=1)
    state_with_2024 = state + [pred_2024]
    
    # Predict 2025 using 2014-2024 data
    pred_2025 = predict_value(historical_data + [pred_2024], years_ahead=1)
    state_full = state_with_2024 + [pred_2025]
    
    for col_idx, value in enumerate(state_full):
        cell = ws.cell(row=row_num, column=col_idx+1)
        cell.value = value
        if col_idx >= 2:  # Numeric columns
            cell.alignment = center_alignment
    row_num += 1

# Union Territories header
ws.cell(row=row_num, column=1).value = ''
ws.cell(row=row_num, column=2).value = 'Union Territories'
for col in range(1, 15):
    cell = ws.cell(row=row_num, column=col)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = center_alignment
row_num += 1

# Union Territories data
ut_data = [
    [1, 'A & N Islands', 5672, 5193, 4897, 4705, 4653, 4497, 4617, 3856, 3949, 4078],
    [2, 'Chandigarh', 29230, 29885, 31624, 32255, 32373, 31342, 23475, 22405, 27169, 27189],
    [3, 'D & N Haveli and Daman & Diu', 6788, 7858, 8979, 9532, 9420, 9731, 11916, 10403, 12491, 12948],
    [3, '', '', '', 3996, 4053, 4046, 4266, '', '', '', ''],  # Daman & Diu sub-row
    [4, 'Delhi', 373693, 374012, 379161, 367046, 362803, 365868, 301645, 271786, 300350, 315087],
    [5, 'Jammu & Kashmir', 154676, 152250, 151935, 152860, 153590, 155486, 161697, 175063, 192857, 288195],
    [6, 'Ladakh', '', '', '', '', '', '', 4020, 4257, 3067, 14340],
    [7, 'Lakshadweep', 658, 819, 644, 823, 820, 947, 1982, 1019, 876, 607],
    [8, 'Puducherry', 14239, 30298, 39393, 40863, 42583, 43810, 28977, 39166, 34987, 33021],
]

for ut in ut_data:
    # Get historical data (columns 2 onwards are the year data)
    historical_data = ut[2:]
    
    # Predict 2024 using 2014-2023 data
    pred_2024 = predict_value(historical_data, years_ahead=1)
    ut_with_2024 = ut + [pred_2024]
    
    # Predict 2025 using 2014-2024 data
    pred_2025 = predict_value(historical_data + [pred_2024], years_ahead=1)
    ut_full = ut_with_2024 + [pred_2025]
    
    for col_idx, value in enumerate(ut_full):
        cell = ws.cell(row=row_num, column=col_idx+1)
        cell.value = value if value != '' else ''
        if col_idx >= 2 and value != '':  # Numeric columns
            cell.alignment = center_alignment
    row_num += 1

# Apply borders to all cells
for row in ws.iter_rows(min_row=1, max_row=row_num-1, min_col=1, max_col=14):
    for cell in row:
        cell.border = border

# Save the workbook
wb.save('c:\\Users\\patel\\Desktop\\hackathon\\UIDIA-Hackathon-26\\births_statewise_and_ut.xlsx')
print("Excel file created successfully: births_statewise_and_ut.xlsx")
